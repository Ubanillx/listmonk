#!/usr/bin/env python3
"""Drive a listmonk marketing workflow with a Bearer integration token.

This script intentionally keeps HTTP and JSON handling in the Python standard
library. Excel parsing uses openpyxl because reliable .xlsx support is the
main requirement for this skill.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any


class APIError(RuntimeError):
    def __init__(self, status: int, message: str, data: Any | None = None) -> None:
        super().__init__(message)
        self.status = status
        self.message = message
        self.data = data


class ExcelDependencyError(RuntimeError):
    pass


class ListmonkClient:
    def __init__(self, base_url: str, bearer_token: str, timeout: int = 30) -> None:
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.headers = {
            "Authorization": f"Bearer {bearer_token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

    def request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        payload: dict[str, Any] | list[Any] | None = None,
    ) -> Any:
        url = f"{self.base_url}{path}"
        if params:
            query = urllib.parse.urlencode(params, doseq=True)
            url = f"{url}?{query}"

        data = None
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")

        req = urllib.request.Request(url, data=data, method=method, headers=self.headers)
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                raw = resp.read()
                if not raw:
                    return None
                if "application/json" in resp.headers.get("Content-Type", ""):
                    parsed = json.loads(raw.decode("utf-8"))
                    return parsed.get("data", parsed)
                return raw.decode("utf-8")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            message = body
            data_out = None
            try:
                parsed = json.loads(body)
                message = parsed.get("message", body)
                data_out = parsed.get("data")
            except json.JSONDecodeError:
                pass
            raise APIError(exc.code, message, data_out) from exc

    def validate_token(self) -> Any:
        return self.request("GET", "/api/lists", params={"minimal": "true", "per_page": "all"})

    def get_list(self, list_id: int) -> dict[str, Any]:
        return self.request("GET", f"/api/lists/{list_id}")

    def query_lists(self, query: str) -> list[dict[str, Any]]:
        page = self.request("GET", "/api/lists", params={"query": query, "page": 1, "per_page": "all"})
        return page.get("results", [])

    def create_list(
        self,
        name: str,
        list_type: str,
        optin: str,
        status: str,
        tags: list[str],
        description: str,
    ) -> dict[str, Any]:
        payload = {
            "name": name,
            "type": list_type,
            "optin": optin,
            "status": status,
            "tags": tags,
            "description": description,
        }
        return self.request("POST", "/api/lists", payload=payload)

    def create_subscriber(self, subscriber: dict[str, Any], list_id: int, preconfirm: bool) -> dict[str, Any]:
        payload = dict(subscriber)
        lists = payload.get("lists", [])
        if list_id not in lists:
            lists = list(lists) + [list_id]
        payload["lists"] = lists
        payload.setdefault("status", "enabled")
        payload["preconfirm_subscriptions"] = preconfirm
        return self.request("POST", "/api/subscribers", payload=payload)

    def query_subscribers(self, search: str, per_page: int | str = "all") -> list[dict[str, Any]]:
        page = self.request("GET", "/api/subscribers", params={"search": search, "page": 1, "per_page": per_page})
        return page.get("results", [])

    def manage_subscriber_lists(
        self,
        subscriber_ids: list[int],
        target_list_ids: list[int],
        status: str,
        action: str = "add",
    ) -> Any:
        payload = {
            "ids": subscriber_ids,
            "action": action,
            "target_list_ids": target_list_ids,
            "status": status,
        }
        return self.request("PUT", "/api/subscribers/lists", payload=payload)

    def clone_template(self, template_id: int, name: str, subject: str | None) -> dict[str, Any]:
        payload: dict[str, Any] = {"name": name}
        if subject:
            payload["subject"] = subject
        return self.request("POST", f"/api/templates/{template_id}/clone", payload=payload)

    def create_campaign(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self.request("POST", "/api/campaigns", payload=payload)

    def update_campaign_status(self, campaign_id: int, status: str) -> dict[str, Any]:
        return self.request("PUT", f"/api/campaigns/{campaign_id}/status", payload={"status": status})

    def get_report_summary(self, campaign_id: int, from_date: str, to_date: str) -> dict[str, Any]:
        return self.request("GET", f"/api/campaigns/{campaign_id}/report/summary", params={"from": from_date, "to": to_date})

    def get_report_timeseries(self, campaign_id: int, from_date: str, to_date: str) -> dict[str, Any]:
        return self.request("GET", f"/api/campaigns/{campaign_id}/report/timeseries", params={"from": from_date, "to": to_date})

    def get_report_links(self, campaign_id: int, from_date: str, to_date: str) -> Any:
        return self.request("GET", f"/api/campaigns/{campaign_id}/report/links", params={"from": from_date, "to": to_date})

    def get_report_recipients(
        self,
        campaign_id: int,
        from_date: str,
        to_date: str,
        per_page: int,
    ) -> dict[str, Any]:
        params = {"from": from_date, "to": to_date, "page": 1, "per_page": per_page}
        return self.request("GET", f"/api/campaigns/{campaign_id}/report/recipients", params=params)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run a listmonk marketing flow with a Bearer token.")
    parser.add_argument("--base-url", required=True, help="Base listmonk URL, for example https://listmonk.example.com")
    parser.add_argument("--bearer-token", required=True, help="listmonk integration Bearer token")

    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--list-id", type=int, help="Existing list ID to reuse")
    target.add_argument("--list-name", help="List name to find or create")

    parser.add_argument("--list-type", default="private", choices=["private", "public"], help="List type when creating a list")
    parser.add_argument("--list-optin", default="single", choices=["single", "double"], help="List opt-in mode when creating a list")
    parser.add_argument("--list-status", default="active", choices=["active", "archived"], help="List status when creating a list")
    parser.add_argument("--list-tags", default="", help="Comma-separated list tags used when creating a list")
    parser.add_argument("--list-description", default="", help="Optional description used when creating a list")

    subscriber_input = parser.add_mutually_exclusive_group(required=True)
    subscriber_input.add_argument("--subscribers-file", help="Path to a JSON file containing an array of subscribers")
    subscriber_input.add_argument("--excel-file", help="Path to a .xlsx file containing subscribers")

    parser.add_argument("--preconfirm-subscriptions", action="store_true", help="Preconfirm subscriber list memberships on create")
    parser.add_argument("--excel-sheet", default="", help="Excel sheet name or 1-based index; defaults to the first sheet")
    parser.add_argument("--email-column", default="", help="Excel e-mail column, by header name or column letter")
    parser.add_argument("--name-column", default="", help="Optional Excel name column, by header name or column letter")
    parser.add_argument("--header-row", type=int, default=1, help="Excel header row number, defaults to 1")
    parser.add_argument("--start-row", type=int, default=0, help="Excel data start row; defaults to header_row + 1")
    parser.add_argument("--skip-empty-rows", action="store_true", help="Skip fully empty Excel rows instead of reporting them")
    parser.add_argument(
        "--dedupe-by-email",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Deduplicate Excel rows by email before creating subscribers",
    )

    parser.add_argument("--source-template-id", type=int, required=True, help="Template ID to clone")
    parser.add_argument("--new-template-name", required=True, help="Name for the cloned template")
    parser.add_argument("--template-subject", default="", help="Optional replacement subject when cloning a tx template")

    parser.add_argument("--campaign-name", required=True, help="Campaign name")
    parser.add_argument("--subject", required=True, help="Campaign subject")
    parser.add_argument("--campaign-body", default="", help="Campaign body, defaults to empty string for template-driven campaigns")
    parser.add_argument("--content-type", default="richtext", choices=["richtext", "html", "markdown", "plain", "visual"])
    parser.add_argument("--messenger", default="email", help="Messenger name, defaults to email")
    parser.add_argument("--from-email", default="", help="Optional campaign from_email")
    parser.add_argument("--daily-send-limit", type=int, required=True, help="Daily send limit for the campaign")
    parser.add_argument("--daily-resume-time", default="09:00", help="Daily resume time in HH:MM local server time")
    parser.add_argument("--send-at", default="", help="Optional RFC3339 send time for scheduling")
    parser.add_argument("--tag", action="append", default=[], help="Campaign tag; pass multiple times for multiple tags")
    parser.add_argument("--attribs-file", help="Path to a JSON object for campaign attribs")

    parser.add_argument("--auto-start", action="store_true", help="Start or schedule the campaign immediately after creation")
    parser.add_argument("--report-from", default="", help="Start date/time for report fetch")
    parser.add_argument("--report-to", default="", help="End date/time for report fetch")
    parser.add_argument("--recipient-per-page", type=int, default=100, help="Recipients page size when fetching recipient analytics")
    parser.add_argument("--verbose", action="store_true", help="Print progress messages to stderr")
    return parser.parse_args()


def load_json_file(path: str, expected: type) -> Any:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(data, expected):
        raise ValueError(f"Expected {expected.__name__} in {path}")
    return data


def log(message: str, *, enabled: bool) -> None:
    if enabled:
        print(message, file=sys.stderr)


def normalize_tags(raw: str) -> list[str]:
    return [item.strip() for item in raw.split(",") if item.strip()]


def normalize_key(value: str) -> str:
    key = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip().lower()).strip("_")
    return key or "column"


def is_column_letter(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z]+", value.strip()))


def extract_emails(raw_value: Any) -> list[str]:
    if raw_value is None:
        return []

    text = str(raw_value).strip()
    if not text:
        return []

    matches = re.findall(r"[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Za-z0-9-]+(?:\.[A-Za-z0-9-]+)+", text)
    out: list[str] = []
    seen: set[str] = set()
    for match in matches:
        email = match.strip()
        key = email.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(email)
    return out


def require_openpyxl() -> tuple[Any, Any, Any]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string, get_column_letter
    except ImportError as exc:
        raise ExcelDependencyError(
            "Excel support requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    return load_workbook, column_index_from_string, get_column_letter


def resolve_sheet(workbook: Any, sheet_spec: str) -> Any:
    if not sheet_spec:
        return workbook[workbook.sheetnames[0]]

    if sheet_spec.isdigit():
        index = int(sheet_spec)
        if index < 1 or index > len(workbook.sheetnames):
            raise ValueError(f"Excel sheet index out of range: {sheet_spec}")
        return workbook[workbook.sheetnames[index - 1]]

    if sheet_spec not in workbook.sheetnames:
        raise ValueError(f"Excel sheet not found: {sheet_spec}")
    return workbook[sheet_spec]


def build_header_maps(header_values: list[Any], max_column: int, get_column_letter: Any) -> tuple[dict[str, int], dict[int, str]]:
    header_to_index: dict[str, int] = {}
    index_to_key: dict[int, str] = {}

    for col_idx in range(1, max_column + 1):
        value = header_values[col_idx - 1] if col_idx - 1 < len(header_values) else None
        if value is None:
            header = get_column_letter(col_idx)
        else:
            header = str(value).strip()
            if not header:
                header = get_column_letter(col_idx)

        header_to_index[header.casefold()] = col_idx
        index_to_key[col_idx] = normalize_key(header)

    return header_to_index, index_to_key


def resolve_column(
    column_spec: str,
    *,
    header_to_index: dict[str, int],
    column_index_from_string: Any,
    label: str,
) -> int:
    if not column_spec:
        raise ValueError(f"{label} is required for Excel import")

    value = column_spec.strip()
    if is_column_letter(value):
        return column_index_from_string(value.upper())

    idx = header_to_index.get(value.casefold())
    if idx is None:
        raise ValueError(f"Could not resolve {label}: {column_spec}")
    return idx


def parse_excel_subscribers(args: argparse.Namespace) -> dict[str, Any]:
    if not args.excel_file:
        return {"source": "json", "subscribers": load_json_file(args.subscribers_file, list), "skipped_rows": [], "failed_rows": []}

    if not args.email_column:
        raise ValueError("--email-column is required when using --excel-file")
    if not str(args.excel_file).lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx Excel files are supported")
    if args.header_row < 1:
        raise ValueError("--header-row must be >= 1")
    if args.start_row < 0:
        raise ValueError("--start-row must be >= 0")

    load_workbook, column_index_from_string, get_column_letter = require_openpyxl()
    workbook = load_workbook(filename=args.excel_file, read_only=True, data_only=True)
    sheet = resolve_sheet(workbook, args.excel_sheet)
    max_column = sheet.max_column or 0
    if max_column < 1:
        raise ValueError("Excel sheet is empty")

    header_rows = sheet.iter_rows(min_row=args.header_row, max_row=args.header_row, values_only=True)
    header_values = list(next(header_rows, ()))
    header_to_index, index_to_key = build_header_maps(header_values, max_column, get_column_letter)
    email_col = resolve_column(
        args.email_column,
        header_to_index=header_to_index,
        column_index_from_string=column_index_from_string,
        label="--email-column",
    )
    name_col = None
    if args.name_column:
        name_col = resolve_column(
            args.name_column,
            header_to_index=header_to_index,
            column_index_from_string=column_index_from_string,
            label="--name-column",
        )

    start_row = args.start_row or (args.header_row + 1)
    seen_emails: set[str] = set()
    subscribers: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []

    row_iter = sheet.iter_rows(min_row=start_row, values_only=True)
    for row_idx, row_values in enumerate(row_iter, start=start_row):
        row_values = tuple(row_values)
        values: dict[int, Any] = {
            col_idx: row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
            for col_idx in range(1, max_column + 1)
        }
        non_empty = any(cell_value not in (None, "") for cell_value in values.values())

        if not non_empty:
            entry = {"row": row_idx, "reason": "empty_row"}
            if args.skip_empty_rows:
                skipped_rows.append(entry)
                continue
            failed_rows.append(entry)
            continue

        raw_email = values.get(email_col)
        raw_email_text = "" if raw_email is None else str(raw_email).strip()
        if not raw_email_text:
            skipped_rows.append({"row": row_idx, "reason": "missing_email"})
            continue
        emails = extract_emails(raw_email_text)
        if not emails:
            failed_rows.append({"row": row_idx, "email": raw_email_text, "reason": "invalid_email"})
            continue

        raw_name = values.get(name_col) if name_col is not None else None
        name = "" if raw_name is None else str(raw_name).strip()

        attribs: dict[str, Any] = {}
        for col_idx in range(1, max_column + 1):
            if col_idx in {email_col, name_col}:
                continue
            value = values.get(col_idx)
            if value in (None, ""):
                continue
            attribs[index_to_key[col_idx]] = value

        for email in emails:
            email_key = email.casefold()
            if args.dedupe_by_email and email_key in seen_emails:
                skipped_rows.append({"row": row_idx, "email": email, "reason": "duplicate_email"})
                continue
            seen_emails.add(email_key)

            subscriber: dict[str, Any] = {"email": email}
            if name:
                subscriber["name"] = name
            if attribs:
                subscriber["attribs"] = dict(attribs)

            subscribers.append({"row": row_idx, "subscriber": subscriber})

    workbook.close()
    return {
        "source": "excel",
        "subscribers": subscribers,
        "skipped_rows": skipped_rows,
        "failed_rows": failed_rows,
    }


def find_or_create_list(client: ListmonkClient, args: argparse.Namespace) -> dict[str, Any]:
    if args.list_id:
        return client.get_list(args.list_id)

    assert args.list_name
    lists = client.query_lists(args.list_name)
    for item in lists:
        if item.get("name") == args.list_name:
            return item

    try:
        return client.create_list(
            name=args.list_name,
            list_type=args.list_type,
            optin=args.list_optin,
            status=args.list_status,
            tags=normalize_tags(args.list_tags),
            description=args.list_description,
        )
    except APIError:
        lists = client.query_lists(args.list_name)
        for item in lists:
            if item.get("name") == args.list_name:
                return item
        raise


def create_subscribers_if_needed(client: ListmonkClient, args: argparse.Namespace, list_id: int) -> dict[str, Any]:
    parsed = parse_excel_subscribers(args)
    source = parsed["source"]
    source_rows = parsed["subscribers"]
    skipped_rows = list(parsed["skipped_rows"])
    failed_rows = list(parsed["failed_rows"])
    created = []

    def reuse_existing_subscriber(subscriber: dict[str, Any]) -> dict[str, Any]:
        email = str(subscriber.get("email", "")).strip()
        matches = client.query_subscribers(email)
        match = next((item for item in matches if str(item.get("email", "")).casefold() == email.casefold()), None)
        if match is None:
            raise APIError(409, "subscriber already exists but could not be queried back")

        desired_status = "confirmed" if args.preconfirm_subscriptions else "unconfirmed"
        lists = match.get("lists", [])
        current = next((item for item in lists if item.get("id") == list_id), None)
        current_status = current.get("subscription_status") if current else None
        if current is None or current_status != desired_status:
            client.manage_subscriber_lists([int(match["id"])], [list_id], desired_status)
            refreshed = client.query_subscribers(email)
            updated = next((item for item in refreshed if str(item.get("email", "")).casefold() == email.casefold()), None)
            if updated is not None:
                match = updated

        return match

    if source == "json":
        for index, subscriber in enumerate(source_rows, start=1):
            if not isinstance(subscriber, dict):
                raise ValueError(f"Each subscriber entry must be a JSON object. Invalid item at index {index}")
            try:
                created.append(client.create_subscriber(subscriber, list_id, args.preconfirm_subscriptions))
            except APIError as exc:
                if exc.status == 409 and subscriber.get("email"):
                    try:
                        created.append(reuse_existing_subscriber(subscriber))
                        continue
                    except APIError as lookup_exc:
                        exc = lookup_exc
                failed_rows.append(
                    {
                        "row": index,
                        "email": subscriber.get("email"),
                        "reason": exc.message,
                        "status": exc.status,
                    }
                )
        return {
            "import_source": source,
            "created_subscribers": created,
            "imported_count": len(created),
            "skipped_rows": skipped_rows,
            "failed_rows": failed_rows,
        }

    for item in source_rows:
        row = int(item["row"])
        subscriber = item["subscriber"]
        try:
            created.append(client.create_subscriber(subscriber, list_id, args.preconfirm_subscriptions))
        except APIError as exc:
            if exc.status == 409 and subscriber.get("email"):
                try:
                    created.append(reuse_existing_subscriber(subscriber))
                    continue
                except APIError as lookup_exc:
                    exc = lookup_exc
            failed_rows.append(
                {
                    "row": row,
                    "email": subscriber.get("email"),
                    "reason": exc.message,
                    "status": exc.status,
                }
            )

    return {
        "import_source": source,
        "created_subscribers": created,
        "imported_count": len(created),
        "skipped_rows": skipped_rows,
        "failed_rows": failed_rows,
    }


def build_campaign_payload(args: argparse.Namespace, list_id: int, template_id: int) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "name": args.campaign_name,
        "subject": args.subject,
        "lists": [list_id],
        "type": "regular",
        "content_type": args.content_type,
        "body": args.campaign_body,
        "template_id": template_id,
        "messenger": args.messenger,
        "daily_send_limit": args.daily_send_limit,
        "daily_resume_time": args.daily_resume_time,
        "tags": args.tag,
    }
    if args.from_email:
        payload["from_email"] = args.from_email
    if args.send_at:
        payload["send_at"] = args.send_at
    if args.attribs_file:
        payload["attribs"] = load_json_file(args.attribs_file, dict)
    return payload


def fetch_reports_if_requested(client: ListmonkClient, args: argparse.Namespace, campaign_id: int) -> dict[str, Any]:
    if not args.report_from or not args.report_to:
        return {}

    reports = {
        "summary": client.get_report_summary(campaign_id, args.report_from, args.report_to),
        "timeseries": client.get_report_timeseries(campaign_id, args.report_from, args.report_to),
        "links": client.get_report_links(campaign_id, args.report_from, args.report_to),
    }

    try:
        reports["recipients"] = client.get_report_recipients(
            campaign_id,
            args.report_from,
            args.report_to,
            args.recipient_per_page,
        )
    except APIError as exc:
        if exc.status == 403:
            reports["recipients_unavailable"] = exc.message
        else:
            raise

    return reports


def main() -> int:
    args = parse_args()
    client = ListmonkClient(args.base_url, args.bearer_token)
    progress: dict[str, Any] = {}

    try:
        log("Validating Bearer token", enabled=args.verbose)
        client.validate_token()

        log("Resolving target list", enabled=args.verbose)
        list_obj = find_or_create_list(client, args)
        list_id = int(list_obj["id"])
        progress["list_id"] = list_id
        progress["list"] = list_obj

        log("Importing subscribers", enabled=args.verbose)
        import_result = create_subscribers_if_needed(client, args, list_id)
        progress.update(import_result)

        log("Cloning template", enabled=args.verbose)
        template = client.clone_template(args.source_template_id, args.new_template_name, args.template_subject or None)
        progress["template_id"] = int(template["id"])
        progress["template"] = template

        log("Creating campaign", enabled=args.verbose)
        campaign = client.create_campaign(build_campaign_payload(args, list_id, int(template["id"])))
        progress["campaign_id"] = int(campaign["id"])
        progress["campaign"] = campaign

        status_result = None
        if args.auto_start:
            log("Starting or scheduling campaign", enabled=args.verbose)
            status_result = client.update_campaign_status(int(campaign["id"]), "running")

        log("Fetching reports", enabled=args.verbose)
        reports = fetch_reports_if_requested(client, args, int(campaign["id"]))

        result = {
            "import_source": progress.get("import_source"),
            "list_id": list_id,
            "template_id": int(template["id"]),
            "campaign_id": int(campaign["id"]),
            "status": (status_result or campaign).get("status", "draft"),
            "list": list_obj,
            "template": template,
            "campaign": campaign,
            "created_subscribers": progress.get("created_subscribers", []),
            "imported_count": progress.get("imported_count", 0),
            "skipped_rows": progress.get("skipped_rows", []),
            "failed_rows": progress.get("failed_rows", []),
        }
        if reports:
            result["reports"] = reports
            if "summary" in reports:
                result["report_summary"] = reports["summary"]

        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return 0
    except (APIError, ExcelDependencyError, ValueError, OSError) as exc:
        if isinstance(exc, APIError):
            error = {"status": exc.status, "message": exc.message, "data": exc.data}
        else:
            error = {"message": str(exc)}
        if progress:
            error["progress"] = progress
        print(json.dumps({"error": error}, ensure_ascii=False, indent=2, default=str), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
