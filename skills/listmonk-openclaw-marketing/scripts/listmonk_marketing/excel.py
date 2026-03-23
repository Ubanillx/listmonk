from __future__ import annotations

import re
from typing import Any


class ExcelDependencyError(RuntimeError):
    pass


def normalize_key(value: str) -> str:
    key = re.sub(r"\W+", "_", value.strip().lower(), flags=re.UNICODE).strip("_")
    return key or "column"


def is_column_letter(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z]+", value.strip()))


def extract_emails(raw_value: Any) -> list[str]:
    if raw_value is None:
        return []

    text = str(raw_value).strip()
    if not text:
        return []

    matches = re.findall(r"[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Za-z0-9-]+(?:\.[A-Za-z0-9-]+)+", text)
    out: list[str] = []
    seen: set[str] = set()
    for match in matches:
        email = match.strip()
        key = email.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(email)
    return out


def require_openpyxl() -> tuple[Any, Any, Any]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string, get_column_letter
    except ImportError as exc:
        raise ExcelDependencyError(
            "Excel support requires openpyxl. Install it with: pip install openpyxl"
        ) from exc

    return load_workbook, column_index_from_string, get_column_letter


def resolve_sheet(workbook: Any, sheet_spec: str) -> Any:
    if not sheet_spec:
        return workbook[workbook.sheetnames[0]]

    if sheet_spec.isdigit():
        index = int(sheet_spec)
        if index < 1 or index > len(workbook.sheetnames):
            raise ValueError(f"Excel sheet index out of range: {sheet_spec}")
        return workbook[workbook.sheetnames[index - 1]]

    if sheet_spec not in workbook.sheetnames:
        raise ValueError(f"Excel sheet not found: {sheet_spec}")
    return workbook[sheet_spec]


def build_header_maps(
    header_values: list[Any],
    max_column: int,
    get_column_letter: Any,
) -> tuple[dict[str, int], dict[int, str]]:
    header_to_index: dict[str, int] = {}
    index_to_key: dict[int, str] = {}

    for col_idx in range(1, max_column + 1):
        value = header_values[col_idx - 1] if col_idx - 1 < len(header_values) else None
        if value is None:
            header = get_column_letter(col_idx)
        else:
            header = str(value).strip()
            if not header:
                header = get_column_letter(col_idx)

        header_to_index[header.casefold()] = col_idx
        index_to_key[col_idx] = normalize_key(header)

    return header_to_index, index_to_key


def resolve_column(
    column_spec: str,
    *,
    header_to_index: dict[str, int],
    column_index_from_string: Any,
    label: str,
) -> int:
    if not column_spec:
        raise ValueError(f"{label} is required for Excel import")

    value = column_spec.strip()
    idx = header_to_index.get(value.casefold())
    if idx is not None:
        return idx

    if is_column_letter(value):
        return column_index_from_string(value.upper())

    raise ValueError(f"Could not resolve {label}: {column_spec}")


def parse_excel_subscribers(
    *,
    excel_file: str,
    email_column: str,
    name_column: str = "",
    excel_sheet: str = "",
    header_row: int = 1,
    start_row: int = 0,
    skip_empty_rows: bool = False,
    dedupe_by_email: bool = True,
) -> dict[str, Any]:
    if not email_column:
        raise ValueError("--email-column is required when using --excel-file")
    if not str(excel_file).lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx Excel files are supported")
    if header_row < 1:
        raise ValueError("--header-row must be >= 1")
    if start_row < 0:
        raise ValueError("--start-row must be >= 0")

    load_workbook, column_index_from_string, get_column_letter = require_openpyxl()
    workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
    try:
        sheet = resolve_sheet(workbook, excel_sheet)
        max_column = sheet.max_column or 0
        if max_column < 1:
            raise ValueError("Excel sheet is empty")

        header_rows = sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        header_values = list(next(header_rows, ()))
        header_to_index, index_to_key = build_header_maps(header_values, max_column, get_column_letter)
        email_col = resolve_column(
            email_column,
            header_to_index=header_to_index,
            column_index_from_string=column_index_from_string,
            label="--email-column",
        )
        name_col = None
        if name_column:
            name_col = resolve_column(
                name_column,
                header_to_index=header_to_index,
                column_index_from_string=column_index_from_string,
                label="--name-column",
            )

        start = start_row or (header_row + 1)
        seen_emails: set[str] = set()
        subscribers: list[dict[str, Any]] = []
        skipped_rows: list[dict[str, Any]] = []
        failed_rows: list[dict[str, Any]] = []

        row_iter = sheet.iter_rows(min_row=start, values_only=True)
        for row_idx, row_values in enumerate(row_iter, start=start):
            row_values = tuple(row_values)
            values: dict[int, Any] = {
                col_idx: row_values[col_idx - 1] if col_idx - 1 < len(row_values) else None
                for col_idx in range(1, max_column + 1)
            }
            non_empty = any(cell_value not in (None, "") for cell_value in values.values())

            if not non_empty:
                entry = {"row": row_idx, "reason": "empty_row"}
                if skip_empty_rows:
                    skipped_rows.append(entry)
                    continue
                failed_rows.append(entry)
                continue

            raw_email = values.get(email_col)
            raw_email_text = "" if raw_email is None else str(raw_email).strip()
            if not raw_email_text:
                skipped_rows.append({"row": row_idx, "reason": "missing_email"})
                continue

            emails = extract_emails(raw_email_text)
            if not emails:
                failed_rows.append({"row": row_idx, "email": raw_email_text, "reason": "invalid_email"})
                continue

            raw_name = values.get(name_col) if name_col is not None else None
            name = "" if raw_name is None else str(raw_name).strip()

            attribs: dict[str, Any] = {}
            for col_idx in range(1, max_column + 1):
                if col_idx in {email_col, name_col}:
                    continue
                value = values.get(col_idx)
                if value in (None, ""):
                    continue
                attribs[index_to_key[col_idx]] = value

            for email in emails:
                email_key = email.casefold()
                if dedupe_by_email and email_key in seen_emails:
                    skipped_rows.append({"row": row_idx, "email": email, "reason": "duplicate_email"})
                    continue
                seen_emails.add(email_key)

                subscriber: dict[str, Any] = {"email": email}
                if name:
                    subscriber["name"] = name
                if attribs:
                    subscriber["attribs"] = dict(attribs)

                subscribers.append({"row": row_idx, "subscriber": subscriber})

        return {
            "source": "excel",
            "subscribers": subscribers,
            "skipped_rows": skipped_rows,
            "failed_rows": failed_rows,
        }
    finally:
        workbook.close()
